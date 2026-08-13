"""
CSV Report Automator
=====================
Takes a messy CSV export (inconsistent casing, extra whitespace, missing
values, duplicate rows) and produces a clean, formatted Excel report with
a summary sheet and a chart — the kind of task small businesses pay for
because it saves them hours of manual spreadsheet cleanup every week.

USAGE
-----
    python csv_report_automator.py input.csv output_report.xlsx

WHAT IT DOES
------------
1. Loads the CSV
2. Cleans it: strips whitespace from headers/values, normalizes text
   casing, drops exact duplicate rows, flags missing values
3. Computes summary stats: totals and averages per numeric column,
   breakdowns by each categorical column
4. Writes a formatted .xlsx with:
   - "Summary" sheet: key totals + a bar chart
   - "Cleaned Data" sheet: the full cleaned dataset

This is a portfolio/demo piece — column names are auto-detected so it
works on most simple tabular CSVs, not just the sample data.
"""

import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
TITLE_FONT = Font(name="Arial", bold=True, size=14)
BODY_FONT = Font(name="Arial", size=11)


def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    # Clean column names
    df.columns = [c.strip() for c in df.columns]

    # Strip whitespace from all string cells, normalize case for object columns
    for col in df.select_dtypes(include=["object", "string"]).columns:
        df[col] = df[col].astype(str).str.strip()
        # Title-case categorical-looking text columns (not free-text notes)
        if df[col].str.len().mean() < 20:  # heuristic: short values = likely categorical
            df[col] = df[col].str.title()

    # Drop exact duplicate rows
    before = len(df)
    df = df.drop_duplicates()
    removed = before - len(df)

    # Convert numeric-looking columns
    for col in df.columns:
        converted = pd.to_numeric(df[col], errors="coerce")
        if converted.notna().sum() >= len(df) * 0.5:  # mostly numeric -> convert
            df[col] = converted

    df.attrs["duplicates_removed"] = removed
    return df


def build_summary(df: pd.DataFrame) -> dict:
    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    categorical_cols = [c for c in df.select_dtypes(include=["object", "string"]).columns
                         if df[c].nunique() <= 20]

    summary = {
        "row_count": len(df),
        "duplicates_removed": df.attrs.get("duplicates_removed", 0),
        "missing_values": {col: int(df[col].isna().sum()) for col in df.columns
                            if df[col].isna().sum() > 0},
        "numeric_totals": {col: round(float(df[col].sum(skipna=True)), 2) for col in numeric_cols},
        "numeric_averages": {col: round(float(df[col].mean(skipna=True)), 2) for col in numeric_cols},
        "category_breakdowns": {},
    }

    for col in categorical_cols:
        summary["category_breakdowns"][col] = df[col].value_counts().to_dict()

    return summary


def write_report(df: pd.DataFrame, summary: dict, output_path: str):
    wb = Workbook()

    # ---------- Summary sheet ----------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Automated Data Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    row = 3
    ws[f"A{row}"] = "Rows processed:"
    ws[f"B{row}"] = summary["row_count"]
    row += 1
    ws[f"A{row}"] = "Duplicate rows removed:"
    ws[f"B{row}"] = summary["duplicates_removed"]
    row += 2

    if summary["missing_values"]:
        ws[f"A{row}"] = "Missing values by column"
        ws[f"A{row}"].font = Font(name="Arial", bold=True)
        row += 1
        for col, count in summary["missing_values"].items():
            ws[f"A{row}"] = col
            ws[f"B{row}"] = count
            row += 1
        row += 1

    ws[f"A{row}"] = "Totals"
    ws[f"A{row}"].font = Font(name="Arial", bold=True)
    row += 1
    totals_start_row = row
    for col, total in summary["numeric_totals"].items():
        ws[f"A{row}"] = col
        ws[f"B{row}"] = total
        row += 1
    totals_end_row = row - 1
    row += 1

    ws[f"A{row}"] = "Averages"
    ws[f"A{row}"].font = Font(name="Arial", bold=True)
    row += 1
    for col, avg in summary["numeric_averages"].items():
        ws[f"A{row}"] = col
        ws[f"B{row}"] = avg
        row += 1
    row += 1

    for col_name, breakdown in summary["category_breakdowns"].items():
        ws[f"A{row}"] = f"Breakdown by {col_name}"
        ws[f"A{row}"].font = Font(name="Arial", bold=True)
        row += 1
        for value, count in breakdown.items():
            ws[f"A{row}"] = value
            ws[f"B{row}"] = count
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16

    # Add a bar chart of totals, if there's at least one numeric column
    if summary["numeric_totals"]:
        chart = BarChart()
        chart.title = "Totals by Column"
        chart.y_axis.title = "Value"
        data = Reference(ws, min_col=2, min_row=totals_start_row, max_row=totals_end_row)
        cats = Reference(ws, min_col=1, min_row=totals_start_row, max_row=totals_end_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        ws.add_chart(chart, f"D3")

    # ---------- Cleaned Data sheet ----------
    ws2 = wb.create_sheet("Cleaned Data")
    for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.font = BODY_FONT

    for col_cells in ws2.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws2.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

    wb.save(output_path)


def main():
    if len(sys.argv) != 3:
        print("Usage: python csv_report_automator.py input.csv output_report.xlsx")
        sys.exit(1)

    input_path, output_path = sys.argv[1], sys.argv[2]

    df = pd.read_csv(input_path)
    df = clean_data(df)
    summary = build_summary(df)
    write_report(df, summary, output_path)

    print(f"Report written to {output_path}")
    print(f"Rows processed: {summary['row_count']}")
    print(f"Duplicates removed: {summary['duplicates_removed']}")


if __name__ == "__main__":
    main()
